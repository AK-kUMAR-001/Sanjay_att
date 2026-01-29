import sqlite3
import os
from datetime import datetime, timedelta
import pandas as pd
from openpyxl.styles import PatternFill, Font

DB_FOLDER = "Attendance"
DB_NAME = "attendance.db"
DB_PATH = os.path.join(DB_FOLDER, DB_NAME)
EXCEL_FILE = os.path.join(DB_FOLDER, "Attendance_Log.xlsx")

def init_db():
    if not os.path.exists(DB_FOLDER):
        os.makedirs(DB_FOLDER)
    
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Create students table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS students (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL
        )
    ''')
    
    # Create attendance table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id TEXT NOT NULL,
            name TEXT NOT NULL,
            date TEXT NOT NULL,
            time TEXT NOT NULL
        )
    ''')
    
    conn.commit()
    conn.close()
    print("Database initialized successfully.")
    init_excel()

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        df = pd.DataFrame(columns=["Student ID", "Name", "Date", "Time", "Status"])
        with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            worksheet = writer.sheets['Sheet1']
            # Format Header
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
        print(f"Created new Excel file: {EXCEL_FILE}")

def log_to_excel(student_id, name, date_str, time_str):
    try:
        new_entry = pd.DataFrame([{
            "Student ID": student_id,
            "Name": name,
            "Date": date_str,
            "Time": time_str,
            "Status": "Present"
        }])
        
        # Append to Excel
        if os.path.exists(EXCEL_FILE):
            with pd.ExcelWriter(EXCEL_FILE, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                # Load existing sheet to find next empty row
                try:
                    writer.book = pd.read_excel(EXCEL_FILE, engine='openpyxl') # Load to check
                    start_row = writer.book.shape[0] + 1
                except:
                    start_row = 0
                
                # We actually just want to append using pandas functionality if possible, 
                # but pandas append is deprecated.
                # Simpler way: Read, Append, Write back (Inefficient for huge data but fine here)
                try:
                     df_existing = pd.read_excel(EXCEL_FILE)
                except ValueError:
                     # Sheet might be empty or file corrupt
                     df_existing = pd.DataFrame(columns=["Student ID", "Name", "Date", "Time", "Status"])

                # Check if this student is already marked within last 3 minutes
                if not df_existing.empty:
                    student_rows = df_existing[df_existing['Student ID'].astype(str) == str(student_id)]
                    if not student_rows.empty:
                        last_row = student_rows.iloc[-1]
                        last_date = str(last_row['Date'])
                        last_time = str(last_row['Time'])
                        try:
                            last_dt = datetime.strptime(f"{last_date} {last_time}", "%Y-%m-%d %H:%M:%S")
                            current_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")
                            if (current_dt - last_dt) < timedelta(minutes=3):
                                 print(f"Skipping Excel log: {name} marked recently.")
                                 return
                        except ValueError:
                            pass

                df_combined = pd.concat([df_existing, new_entry], ignore_index=True)
                
                with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='w') as writer_new:
                     df_combined.to_excel(writer_new, index=False)
                     # Re-apply header styles
                     worksheet = writer_new.sheets['Sheet1']
                     header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                     header_font = Font(color="FFFFFF", bold=True)
                     for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                     
    except Exception as e:
        print(f"Error logging to Excel: {e}")

        print(f"Logged to Excel: {name}")
    except Exception as e:
        print(f"Error logging to Excel: {e}")

def mark_attendance(student_id, name):
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")
    
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    try:
        # Check if already marked within the last 3 minutes
        cursor.execute("SELECT date, time FROM attendance WHERE student_id = ? ORDER BY id DESC LIMIT 1", (student_id,))
        last_record = cursor.fetchone()
        
        if last_record:
            last_date_str, last_time_str = last_record
            try:
                last_dt = datetime.strptime(f"{last_date_str} {last_time_str}", "%Y-%m-%d %H:%M:%S")
                if (now - last_dt) < timedelta(minutes=3):
                    print(f"Attendance already marked for {name} recently ({now - last_dt}).")
                    return
            except ValueError:
                pass # Ignore parsing errors, just mark new attendance

        cursor.execute('''
            INSERT INTO attendance (student_id, name, date, time)
            VALUES (?, ?, ?, ?)
        ''', (student_id, name, date_str, time_str))
        conn.commit()
        print(f"Marked attendance for {name} ({student_id}) at {time_str}")
        
        # Append to Excel Log (Check for duplicates first)
        log_to_excel(student_id, name, date_str, time_str)
        
    except sqlite3.IntegrityError:
        # print(f"Attendance already marked for {name} today.")
        pass
    finally:
        conn.close()

def get_attendance_records():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM attendance ORDER BY date DESC, time DESC")
    records = cursor.fetchall()
    conn.close()
    return records

def generate_session_report(session_start_time, session_name="Session"):
    """
    Generates a formatted Excel report for the current session (records > session_start_time).
    Returns: (file_path, summary_text, unique_attendees_list)
    """
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Filter by date and time
    start_date = session_start_time.strftime("%Y-%m-%d")
    start_time_str = session_start_time.strftime("%H:%M:%S")
    
    # Get records from DB that match the date and are after the start time
    # Note: This simple comparison works if the session doesn't cross midnight
    cursor.execute('''
        SELECT student_id, name, date, time 
        FROM attendance 
        WHERE date = ? AND time >= ?
        ORDER BY time ASC
    ''', (start_date, start_time_str))
    
    rows = cursor.fetchall()
    conn.close()
    
    if not rows:
        return None, "No attendance recorded in this session.", []
        
    # Filter distinct students (keep first occurrence per session)
    seen_ids = set()
    unique_rows = []
    unique_attendees = [] # List of (name, id, time) for WhatsApp
    
    for row in rows:
        student_id = row[0]
        if student_id not in seen_ids:
            seen_ids.add(student_id)
            unique_rows.append(row)
            unique_attendees.append((row[1], row[0], row[3])) # name, id, time
            
    # Create DataFrame
    df = pd.DataFrame(unique_rows, columns=["Student ID", "Name", "Date", "Time"])
    df["Status"] = "Present"
    df["Session"] = session_name
    
    # Generate Filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    clean_session_name = "".join(c for c in session_name if c.isalnum() or c in (' ', '_', '-')).strip().replace(' ', '_')
    filename = f"Attendance_{clean_session_name}_{timestamp}.xlsx"
    file_path = os.path.join(DB_FOLDER, filename)
    
    # Write to Excel with Formatting
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Report')
        worksheet = writer.sheets['Report']
        
        # Styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Blue
        header_font = Font(color="FFFFFF", bold=True, size=12)
        row_font = Font(size=11)
        
        # Apply Header Style
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            
        # Auto-adjust column width
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
    summary = f"Session Report: {len(unique_rows)} students present.\nSaved to: {filename}"
    return file_path, summary, unique_attendees


if __name__ == "__main__":
    init_db()
